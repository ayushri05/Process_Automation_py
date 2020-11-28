from datetime import date
import openpyxl as xl
#from editpyxl import Workbook
import csv
import codecs

#exampleFile = open('RatePlanCharge.csv')
#list(csv.reader(open('RatePlanCharge.csv', 'rb'), delimiter='\t'))
#ws1 = csv.reader(codecs.open('RatePlanCharge.csv', 'rb', 'utf-16'))


def handel_XL():

    wb = Workbook()

    ws1 = csv.reader(codecs.open('RatePlanCharge.csv',
                                 'rU', 'utf-8'))
    wb1 = list(ws1)
    #c = wb1[1][1]
    # print(c)
    # exampleData
    # print(wb1)

    #filename = "RatePlanCharge.csv"
    #wb1 = xl.load_workbook(filename)
    #ws1 = wb1.worksheets[0]

    filename1 = "Excel_Automation.xlsm"
    wb2 = wb.open(filename1)

    #pipwb2 = xl.load_workbook(filename1, keep_vba=True)
    ws2 = wb2['part1 zuora>enc data']

   # print(ws2.cell(row=34151, column=13).value)
    mr = 0
    mc = ws2.max_column
   # cells = ws2.columns['M']
    # print()

    for max_row, row in enumerate(ws2, 1):
        mr = mr+1
        if all(c.value is None for c in row):
            break
    #csvRows = []
    # for col in range(1, k-1):
    #   print(ws2.cell(row=mr, column=col).value)
    #  csvRows.append(ws2.cell(row=mr, column=col).value)
    print(mr)
    total_Row_Csv = len(wb1)
    #total_row = len(ws2)
    # print(total_row)
    # print(total_Row_Csv)
    # copying the cell values from source
    # excel file to destination excel file

    mc = 17
    k = 12
    a = b = -1
    ws2.cell(row=mr+1, column=k).value = date.today()

    for i in range(mr, mr + total_Row_Csv):
        a = a+1
        b = 0
        for j in range(1, mc+1):
            # reading cell value from source excel file
            if j < k:
               # print(ws2.cell(row=i-1, column=j).value)
                ws2.cell(row=i, column=j).value = ws2.cell(
                    row=i-1, column=j).value
                #print(i, j, k)
                # print(j)
                #print(ws2.cell(row=i, column=j).value)
            elif j > k:
                #print(a, b)
               # print(wb1[a][b])
                c = wb1[a][b]

                b = b+1
                #print('In Loop')
                # writing the read value to destination excel file
                ws2.cell(row=i, column=j).value = c
            #print(ws2.cell(row=i, column=j).value)

    #ws2.cell(row=i, column=1).value = csvRows
    wb2.save('Excel_Automation_05-10-2020.xlsm')


handel_XL()
